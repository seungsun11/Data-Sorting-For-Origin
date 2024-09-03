from openpyxl import load_workbook, Workbook
import pandas as pd
import os

def convert_xls_to_xlsx(xls_path, xlsx_path):
    # .xls 파일 읽기
    xls = pd.ExcelFile(xls_path)
    
    # 새로운 엑셀 라이터 객체 생성
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        for sheet_name in xls.sheet_names:
            # 각 시트를 데이터프레임으로 읽기
            df = pd.read_excel(xls_path, sheet_name=sheet_name)
            
            # 데이터프레임을 새로운 .xlsx 파일에 쓰기
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def apply_absolute_values(file_path, new_file_path):
    
    convert_xls_to_xlsx(file_path, new_file_path)
    
    wb = load_workbook(new_file_path)
    try:
        wb.remove(wb['Calc'])
        wb.remove(wb['Settings'])
    except:
        pass

    # 모든 시트를 처리
    for sheet in wb.worksheets:
        columns_to_process = []
        # 1행과 2행을 확인하여 'AI'가 있는 열을 찾기
        for col in range(1, sheet.max_column + 1):
            cell_value_1 = sheet.cell(row=1, column=col).value
            cell_value_2 = sheet.cell(row=2, column=col).value
            if cell_value_1 == 'AI' or cell_value_2 == 'AI':
                columns_to_process.append(col)
        
        # 찾은 열의 값들을 절대값으로 변환
        for col in columns_to_process:
            for row in range(2, sheet.max_row + 1):  # 데이터는 2행부터 시작한다고 가정
                cell = sheet.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.value = abs(cell.value)
        
    wb.save(new_file_path)

#endregion

dir_path ="C:\\Users\\user\\Desktop\\측정 데이터\\KHM\\240514 BN"

file = "example.xls"
new_file = "output.xlsx"

apply_absolute_values(dir_path + "\Hope.xls", dir_path + "\Hope.xlsx")