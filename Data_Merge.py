from openpyxl import load_workbook, Workbook
import pandas as pd
import os
def convert_xls_to_xlsx(xls_path, xlsx_path):
    # .xls 파일 읽기
    xls = pd.ExcelFile(xls_path)
    
    # 새로운 엑셀 라이터 객체 생성
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        for sheet_name in xls.sheet_names:
            # 각 시트를 데이터프레임으로 읽기
            df = pd.read_excel(xls_path, sheet_name=sheet_name)
            
            # 데이터프레임을 새로운 .xlsx 파일에 쓰기
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def Data_Merge_file(original_file_path, merged_file_path):
    #region 

    convert_xls_to_xlsx(original_file_path, 'test.xlsx')
    file_path = 'test.xlsx'

    # 엑셀 파일 읽기
    wb = load_workbook(file_path)
    try:
        wb.remove(wb['Calc'])
        wb.remove(wb['Settings'])
    except:
        pass


    # 새로운 워크북 생성
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "MergedSheet"

    # 각 시트의 데이터를 담을 리스트 초기화
    column_offset = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 시트 이름을 추가 (헤더로 추가)
        merged_ws.cell(row=1, column=column_offset + 1, value=f"Sheet: {sheet_name}")
        
        # 각 시트의 데이터를 새로운 열에 추가
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=2):
            for col_idx, cell_value in enumerate(row, start=1):
                # 1열과 2열을 바꾸기
                if col_idx == 1:
                    merged_ws.cell(row=row_idx, column=column_offset + 2, value=cell_value)
                elif col_idx == 2:
                    merged_ws.cell(row=row_idx, column=column_offset + 1, value=cell_value)
                elif col_idx ==3:
                    if cell_value == '#REF':
                        merged_ws.cell(row=row_idx, column=column_offset + col_idx, value=0)
                    else:
                        merged_ws.cell(row=row_idx, column=column_offset + col_idx, value=cell_value)
        
        # 다음 시트 데이터를 위한 열 오프셋 조정
        column_offset += ws.max_column
        
        
   


# 결합된 데이터를 새로운 엑셀 파일로 저장
    merged_wb.save(merged_file_path)

    os.remove(original_file_path+'test.xlsx')

#endregion